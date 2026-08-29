"""多格式结构化解析器（docx / xlsx / markdown）。

统一输出 `ParseBlock`（block_type + text + page + reading_order），
保留各格式固有结构：
- docx：标题层级（heading，text 带 # 前缀表层级）+ 段落 + 表格（table 原子块）
- xlsx：工作表名（section）+ 单元格行列数据（table 原子块，保留行列分隔）
- md  ：标题 / 列表 / 表格 / 代码块，**保留 markdown 语法标记**

所有解析器对损坏文件抛 `UnsupportedFormatError`（带友好中文提示），
由上层 parse 路由捕获后返回 ok=false 信封。
"""
from __future__ import annotations

import io
import re

from .cleaner import estimate_tokens
from .models import ParseBlock


class UnsupportedFormatError(Exception):
    """文件损坏或格式不受支持时的友好错误。"""


# ---------------------------------------------------------------------------
# docx
# ---------------------------------------------------------------------------
def parse_docx(data: bytes, do_clean: bool = True) -> list[ParseBlock]:
    try:
        import docx
        from docx.table import Table
        from docx.text.paragraph import Paragraph
        from docx.oxml.table import CT_Tbl
        from docx.oxml.text.paragraph import CT_P
    except ImportError:
        raise UnsupportedFormatError("未安装 python-docx，无法解析 Word 文档") from None

    try:
        document = docx.Document(io.BytesIO(data))
    except Exception as e:
        raise UnsupportedFormatError(f"无法解析 Word 文档，文件可能已损坏或格式不正确：{e}") from e

    blocks: list[ParseBlock] = []
    order = 0

    def add(block_type: str, text: str):
        nonlocal order
        t = text.strip()
        if not t:
            return
        blocks.append(ParseBlock(
            block_type=block_type, text=t, bbox=None, page=1,
            reading_order=order, token_count=estimate_tokens(t)))
        order += 1

    # 按文档真实顺序遍历段落与表格（保持阅读顺序）
    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            para = Paragraph(child, document)
            text = para.text
            if not text.strip():
                continue
            level = _heading_level(para.style.name if para.style else "")
            if level is not None:
                # 标题：text 带 # 前缀表层级，block_type=heading（chunker 维护 heading_path）
                add("heading", "#" * level + " " + text.strip())
            else:
                add("text", text)
        elif isinstance(child, CT_Tbl):
            table = Table(child, document)
            rows = []
            for row in table.rows:
                cells = [c.text.strip().replace("\n", " ") for c in row.cells]
                rows.append(" | ".join(cells))
            if rows:
                add("table", "\n".join(rows))
    return blocks


def _heading_level(style_name: str) -> int | None:
    """从 docx 段落样式名提取标题层级；非标题返回 None。"""
    if not style_name:
        return None
    s = style_name.lower()
    m = re.search(r"(heading|标题)\s*(\d+)", s)
    if m:
        return int(m.group(2))
    if s.startswith("title") or s == "标题":
        return 1
    return None


# ---------------------------------------------------------------------------
# xlsx
# ---------------------------------------------------------------------------
def parse_xlsx(data: bytes, do_clean: bool = True) -> list[ParseBlock]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise UnsupportedFormatError("未安装 openpyxl，无法解析 Excel 表格") from None

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise UnsupportedFormatError(f"无法解析 Excel 表格，文件可能已损坏或格式不正确：{e}") from e

    blocks: list[ParseBlock] = []
    order = 0
    page = 1

    def add(block_type: str, text: str, p: int):
        nonlocal order
        t = text.strip()
        if not t:
            return
        blocks.append(ParseBlock(
            block_type=block_type, text=t, bbox=None, page=p,
            reading_order=order, token_count=estimate_tokens(t)))
        order += 1

    for ws in wb.worksheets:
        sheet_name = ws.title or f"Sheet{page}"
        add("section", "# " + sheet_name, page)
        rows = []
        for row in ws.iter_rows(values_only=True):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            cells = ["" if c is None else str(c).strip().replace("\n", " ") for c in row]
            rows.append(" | ".join(cells))
        if rows:
            add("table", "\n".join(rows), page)
        page += 1
    wb.close()
    return blocks


# ---------------------------------------------------------------------------
# markdown
# ---------------------------------------------------------------------------
_HEADING_RE = re.compile(r"^(#{1,6})\s+(.*)$")
_LIST_RE = re.compile(r"^(\s*)([-*+]|\d+\.)\s+(.*)$")
_TABLE_RE = re.compile(r"^\s*\|.*\|\s*$")
_QUOTE_RE = re.compile(r"^(\s*>\s?)(.*)$")


def parse_markdown(data: bytes, do_clean: bool = True) -> list[ParseBlock]:
    text = data.decode("utf-8", errors="ignore")
    lines = text.split("\n")

    blocks: list[ParseBlock] = []
    order = 0
    cur_type = "text"
    cur_lines: list[str] = []
    in_code = False

    def flush():
        nonlocal order, cur_type, cur_lines
        body = "\n".join(cur_lines).strip()
        if body:
            blocks.append(ParseBlock(
                block_type=cur_type, text=body, bbox=None, page=1,
                reading_order=order, token_count=estimate_tokens(body)))
            order += 1
        cur_lines = []

    def switch(new_type: str):
        nonlocal cur_type
        flush()
        cur_type = new_type

    for raw in lines:
        line = raw.rstrip("\r")
        stripped = line.strip()

        # 代码块围栏
        if stripped.startswith("```"):
            if in_code:
                cur_lines.append(line)
                switch("text")  # 代码块结束，flush
                in_code = False
            else:
                switch("text")
                cur_type = "text"
                cur_lines.append(line)  # 保留围栏标记
                in_code = True
            continue

        if in_code:
            cur_lines.append(line)
            continue

        # 空行 → 分段
        if not stripped:
            flush()
            continue

        # 标题
        m = _HEADING_RE.match(line)
        if m:
            switch("heading")
            cur_lines.append(line)  # 保留 # 标记
            flush()
            continue

        # 表格
        if _TABLE_RE.match(line):
            if cur_type != "table":
                switch("table")
            cur_lines.append(line)
            continue

        # 列表
        m = _LIST_RE.match(line)
        if m:
            if cur_type != "list":
                switch("list")
            cur_lines.append(line)
            continue

        # 引用
        if _QUOTE_RE.match(line):
            if cur_type != "text":
                switch("text")
            cur_lines.append(line)
            continue

        # 普通段落
        if cur_type not in ("text", "list", "quote"):
            switch("text")
        cur_lines.append(line)

    flush()
    return blocks
